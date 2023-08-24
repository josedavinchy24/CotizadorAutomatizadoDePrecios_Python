from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

archivoExcel = openpyxl.load_workbook('Archivo insumo.xlsx')                                            #SE ELIGE EL ARCHIVO DE EXCEL QUE SE VA USAR 

hoja = archivoExcel.active                                                                              #SE ESCOGE LA HOJA DE EXCEL

celdas = hoja['A2' : 'E7']                                                                              #SE ASIGNA EL RANGO QUE SE VA A UTILIZAR 

listaproductos= []                                                                                    
contfila = 1

for fila in celdas:
    producto = [celda.value for celda in fila]                                                          
    listaproductos.append(producto)                                                                     #SE GUARDAN LOS VALORES DE LAS CELDAS EN EL ARREGLO


for producto in listaproductos:
    contfila += 1
    driver = webdriver.Chrome("./chromedriver.exe")
    driver.get("https://www.mercadolibre.com.co/")
    buscador = driver.find_element_by_class_name("nav-search-input")                                    #TRAER BUSCADOR CON LA CLASE 
    buscador.clear()                                                                                    #LIMPIAR BUSCADOR
    nombreproducto=producto[0]                                                                          #NOMBRE DEL PRODUCTO
    buscador.send_keys(nombreproducto)                                                                  #BUSCAR 
    buscador.send_keys(Keys.RETURN)                                                                     #OBTENER RESULTADO DE LA PAGINA
    cantidad = driver.find_elements_by_xpath("//h2[@class='ui-search-item__title']")                    #OBTENER TITULOS DE LOS RESULTADOS
    guardarCantidad = 0                                                                                 
    for i in cantidad:                                                                                  #SE CUENTA LA CANTIDAD DE PRODUCTOS
        guardarCantidad += 1
    
    if guardarCantidad != 0:                                                                            #REVISAR SI HUBO RESULTADOS 

        #SE OBTIENE EL PRECIO DE LOS PRODUCTOS
        precioProductos = driver.find_elements_by_xpath("//li[@class='ui-search-layout__item']//div[@class='ui-search-result__wrapper']//div[@class='andes-card andes-card--flat andes-card--default ui-search-result ui-search-result--core andes-card--padding-default']//div[@class='ui-search-result__content-wrapper']//div[@class='ui-search-result__content-columns']//div[@class='ui-search-result__content-column ui-search-result__content-column--left']//div[@class='ui-search-item__group ui-search-item__group--price']//div[@class='ui-search-item__group__element ui-search-price__part-without-link']//div[@class='ui-search-price ui-search-price--size-medium']//div[@class='ui-search-price__second-line']//span[@class='price-tag ui-search-price__part']//span[@class='price-tag-amount']//span[@class='price-tag-fraction']")
        precioProductos = [ precio.text for precio in precioProductos]
        menor = float(precioProductos[0].replace(".", ""))                                              #PRIMER VALOR PARA REALIZAR LA COMPARACION DE PRECIOS
        

        #SE OBTIENE LOS LINKS DE LOS PRODUCTOS
        linkProductos = driver.find_elements_by_xpath("//div[@class='ui-search-item__group ui-search-item__group--title']//a[1]")
        linkProductos = [ link.get_attribute("href") for link in linkProductos]

        contadorLink = 0                                                                                 # INICIALIZAR CONTADORES PARA GUARDAR POSICION DEL LINK                                                                    
        guardarLink= 0


        for i in precioProductos:                                                                        #SE OBTIENE EL MENOR PRECIO
            convertir=i.replace(".", "")                                                                 #SE ELIMINAN TODOS LOS PUNTOS DEL STRING PARA CONVERTIRLO A FLOAT
            comparar = float(convertir)                                                                 
            if comparar < menor:
                menor = comparar
                guardarLink = contadorLink                                                               #SE GUARDA LA UBICACION DEL PRODUCTO PARA OBTENER EL LINK
            contadorLink += 1
        linkProd = linkProductos[guardarLink]
        
        valorcantidad = hoja.cell(row= contfila ,column = 3)                                             #SE ASIGNA EL VALOR A LA FILA Y COLUMNA DE EXCEL
        valorcantidad.value = guardarCantidad

        valorPrecio = hoja.cell(row= contfila ,column = 4)
        valorPrecio.value = linkProd

        valorLink = hoja.cell(row= contfila ,column = 5)
        valorLink.value = menor

        estado = hoja.cell(row= contfila ,column = 2)
        estado.value = "EXITOSO"

        archivoExcel.save("Archivo insumo.xlsx")                                                         #SE GUARDAN LOS VALORES EN EL EXCEL 
        time.sleep(1)
        driver.close()
    else:
        valorcantidad = hoja.cell(row= contfila ,column = 3)                                             #SE ASIGNA EL VALOR A LA FILA Y COLUMNA DE EXCEL, BUSQUEDA NO ENCONTRADA
        valorcantidad.value = guardarCantidad

        valorPrecio = hoja.cell(row= contfila ,column = 4)
        valorPrecio.value = " "

        valorLink = hoja.cell(row= contfila ,column = 5)
        valorLink.value = " "

        estado = hoja.cell(row= contfila ,column = 2)
        estado.value = "EXITOSO"

        archivoExcel.save("Archivo insumo.xlsx")                                                          #SE GUARDAN LOS VALORES EN EL EXCEL 
        time.sleep(1)
        driver.close()